from pathlib import Path

import pytest

from src.agents.llm_client import (
    LLMRequestError,
    _profile_url,
    model_api_key_configured,
    model_options_payload,
    request_json_design,
    request_text_completion,
    resolve_model_profile,
)
from src.core.config import load_settings


def test_custom_chat_completion_uses_configured_gateway(monkeypatch, tmp_path: Path) -> None:
    config_path = tmp_path / "config.yaml"
    config_path.write_text(
        f"""
models:
  chat_model: "gpt-5-4-2026-03-05"
  embedding_provider: "hashing"
openai:
  provider: "custom_chat_completions"
  api_key_env: "AI_GATEWAY_API_KEY"
  api_key_header: "X-API-Key"
  api_key_prefix: ""
  base_url: "https://iapi-test.proj.com/gpt/v2"
  chat_completions_path: "/gpt-5-4-2026-03-05/chat/completions"
  embedding_enabled: false
paths:
  docs_dir: "{(tmp_path / "docs").as_posix()}"
  jars_dir: "{(tmp_path / "jars").as_posix()}"
  output_dir: "{(tmp_path / "output").as_posix()}"
  vector_store: "{(tmp_path / "output" / "vectors.sqlite3").as_posix()}"
""",
        encoding="utf-8",
    )
    settings = load_settings(config_path)
    monkeypatch.setenv("AI_GATEWAY_API_KEY", "unit-test-secret")
    captured = {}

    class FakeResponse:
        def raise_for_status(self) -> None:
            return None

        def json(self) -> dict:
            return {
                "choices": [
                    {
                        "message": {
                            "content": '{"process_id":"sample","nodes":[],"flows":[],"forms":[]}'
                        }
                    }
                ]
            }

    def fake_post(url, headers, json, timeout, verify=True):
        captured.update({"url": url, "headers": headers, "json": json, "timeout": timeout, "verify": verify})
        return FakeResponse()

    monkeypatch.setattr("src.agents.llm_client.requests.post", fake_post)

    design = request_json_design(settings, "Return JSON only")

    assert design == {"process_id": "sample", "nodes": [], "flows": [], "forms": []}
    assert captured["url"] == "https://iapi-test.proj.com/gpt/v2/gpt-5-4-2026-03-05/chat/completions"
    assert captured["headers"]["X-API-Key"] == "unit-test-secret"
    assert captured["json"]["model"] == "gpt-5-4-2026-03-05"
    assert captured["json"]["messages"][0]["content"] == "Return JSON only"
    assert captured["verify"] is True


def test_direct_openai_chat_completion_profile_uses_bearer_and_max_tokens(monkeypatch, tmp_path: Path) -> None:
    config_path = tmp_path / "config.yaml"
    config_path.write_text(
        f"""
models:
  chat_model: "openai-gpt-4-1-nano-direct"
  embedding_provider: "hashing"
  available_chat_models:
    - id: "openai-gpt-4-1-nano-direct"
      label: "ChatGPT GPT-4.1 Nano (Direct OpenAI)"
      provider: "openai_chat_completions"
      model: "gpt-4.1-nano"
      base_url: "https://api.openai.com"
      chat_completions_path: "/v1/chat/completions"
      api_key_env: "OPENAI_API_KEY"
      api_key_header: "Authorization"
      api_key_prefix: "Bearer "
      verify_ssl: false
      max_output_tokens: 1000
openai:
  provider: "custom_chat_completions"
  api_key_env: "AI_GATEWAY_API_KEY"
  api_key_header: "X-API-Key"
  base_url: "https://iapi-test.proj.com/gpt/v2"
  chat_completions_path: "/gpt-5-4-2026-03-05/chat/completions"
paths:
  docs_dir: "{(tmp_path / "docs").as_posix()}"
  jars_dir: "{(tmp_path / "jars").as_posix()}"
  output_dir: "{(tmp_path / "output").as_posix()}"
  vector_store: "{(tmp_path / "output" / "vectors.sqlite3").as_posix()}"
""",
        encoding="utf-8",
    )
    settings = load_settings(config_path)
    monkeypatch.setenv("OPENAI_API_KEY", "unit-test-openai-key")
    captured = {}

    class FakeResponse:
        def raise_for_status(self) -> None:
            return None

        def json(self) -> dict:
            return {
                "choices": [{"message": {"content": '{"ok": true}'}}],
                "usage": {"prompt_tokens": 8, "completion_tokens": 4, "total_tokens": 12},
            }

    def fake_post(url, headers, json, timeout, verify=True):
        captured.update({"url": url, "headers": headers, "json": json, "timeout": timeout, "verify": verify})
        return FakeResponse()

    monkeypatch.setattr("src.agents.llm_client.requests.post", fake_post)

    text = request_text_completion(settings, "Return JSON only", model_id="openai-gpt-4-1-nano-direct", action="json_design")

    assert text == '{"ok": true}'
    assert captured["url"] == "https://api.openai.com/v1/chat/completions"
    assert captured["headers"]["Authorization"] == "Bearer unit-test-openai-key"
    assert captured["json"]["model"] == "gpt-4.1-nano"
    assert captured["json"]["max_tokens"] == 1000
    assert "max_completion_tokens" not in captured["json"]
    assert captured["json"]["response_format"] == {"type": "json_object"}
    assert captured["verify"] is False


def test_model_key_resolution_does_not_reuse_openai_key_for_other_profiles(monkeypatch, tmp_path: Path) -> None:
    config_path = tmp_path / "config.yaml"
    config_path.write_text(
        f"""
models:
  chat_model: "openai-gpt-4-1-nano-direct"
  available_chat_models:
    - id: "openai-gpt-4-1-nano-direct"
      label: "Direct OpenAI"
      provider: "openai_chat_completions"
      model: "gpt-4.1-nano"
      base_url: "https://api.openai.com"
      chat_completions_path: "/v1/chat/completions"
      api_key_env: "OPENAI_API_KEY"
    - id: "claude-opus-4-6"
      label: "Claude Opus 4.6"
      provider: "custom_messages"
      model: "claude-opus-4-6"
      base_url: "https://example.invalid"
      chat_completions_path: "/claude-opus-4-6-v1"
      api_key_env: "CLAUDE_API_KEY"
openai:
  api_key_env: "AI_GATEWAY_API_KEY"
paths:
  docs_dir: "{(tmp_path / "docs").as_posix()}"
  jars_dir: "{(tmp_path / "jars").as_posix()}"
  output_dir: "{(tmp_path / "output").as_posix()}"
  vector_store: "{(tmp_path / "output" / "vectors.sqlite3").as_posix()}"
""",
        encoding="utf-8",
    )
    settings = load_settings(config_path)
    monkeypatch.setenv("OPENAI_API_KEY", "unit-test-openai-key")
    monkeypatch.delenv("CLAUDE_API_KEY", raising=False)

    assert model_api_key_configured(settings, "openai-gpt-4-1-nano-direct") is True
    assert model_api_key_configured(settings, "claude-opus-4-6") is False


def test_shared_yaml_api_key_is_reused_for_all_model_profiles(monkeypatch, tmp_path: Path) -> None:
    config_path = tmp_path / "config.yaml"
    config_path.write_text(
        f"""
models:
  chat_model: "claude-opus-4-6"
  api_key: "enterprise-shared-key"
  available_chat_models:
    - id: "openai-gpt-4-1-nano-direct"
      label: "Direct OpenAI"
      provider: "openai_chat_completions"
      model: "gpt-4.1-nano"
      base_url: "https://api.openai.com"
      chat_completions_path: "/v1/chat/completions"
      api_key_env: "OPENAI_API_KEY"
      api_key_header: "Authorization"
      api_key_prefix: "Bearer "
    - id: "claude-opus-4-6"
      label: "Claude Opus 4.6"
      provider: "custom_messages"
      model: "claude-opus-4-6"
      base_url: "https://example.invalid"
      chat_completions_path: "/claude-opus-4-6-v1"
      api_key_env: "CLAUDE_API_KEY"
      api_key_header: "X-API-Key"
openai:
  api_key_env: "AI_GATEWAY_API_KEY"
paths:
  docs_dir: "{(tmp_path / "docs").as_posix()}"
  jars_dir: "{(tmp_path / "jars").as_posix()}"
  output_dir: "{(tmp_path / "output").as_posix()}"
  vector_store: "{(tmp_path / "output" / "vectors.sqlite3").as_posix()}"
""",
        encoding="utf-8",
    )
    settings = load_settings(config_path)
    monkeypatch.delenv("OPENAI_API_KEY", raising=False)
    monkeypatch.delenv("CLAUDE_API_KEY", raising=False)
    monkeypatch.delenv("AI_GATEWAY_API_KEY", raising=False)
    captured = {}

    class FakeResponse:
        def raise_for_status(self) -> None:
            return None

        def json(self) -> dict:
            return {"choices": [{"message": {"content": "ok"}}]}

    def fake_post(url, headers, json, timeout, verify=True):
        captured.setdefault("calls", []).append({"url": url, "headers": headers, "json": json})
        return FakeResponse()

    monkeypatch.setattr("src.agents.llm_client.requests.post", fake_post)

    assert model_api_key_configured(settings, "openai-gpt-4-1-nano-direct") is True
    assert model_api_key_configured(settings, "claude-opus-4-6") is True
    assert request_text_completion(settings, "Say OK", model_id="openai-gpt-4-1-nano-direct") == "ok"
    assert request_text_completion(settings, "Say OK", model_id="claude-opus-4-6") == "ok"
    assert captured["calls"][0]["headers"]["Authorization"] == "Bearer enterprise-shared-key"
    assert captured["calls"][1]["headers"]["X-API-Key"] == "enterprise-shared-key"


def test_configured_chat_model_selects_matching_profile_not_first_option(tmp_path: Path) -> None:
    config_path = tmp_path / "config.yaml"
    config_path.write_text(
        f"""
models:
  chat_model: "gpt-5-4-2026-03-05"
  available_chat_models:
    - id: "openai-gpt-4-1-nano-direct"
      label: "Direct OpenAI"
      provider: "openai_chat_completions"
      model: "gpt-4.1-nano"
      base_url: "https://api.openai.com"
      chat_completions_path: "/v1/chat/completions"
      api_key_env: "OPENAI_API_KEY"
    - id: "openai-gpt-5-4"
      label: "OpenAI GPT-5.4"
      provider: "custom_chat_completions"
      model: "gpt-5-4-2026-03-05"
      base_url: "https://iapi-test.proj.com/gpt/v2"
      chat_completions_path: "/gpt-5-4-2026-03-05/chat/completions"
      api_key_env: "AI_GATEWAY_API_KEY"
openai:
  api_key_env: "AI_GATEWAY_API_KEY"
paths:
  docs_dir: "{(tmp_path / "docs").as_posix()}"
  jars_dir: "{(tmp_path / "jars").as_posix()}"
  output_dir: "{(tmp_path / "output").as_posix()}"
  vector_store: "{(tmp_path / "output" / "vectors.sqlite3").as_posix()}"
""",
        encoding="utf-8",
    )
    settings = load_settings(config_path)

    assert resolve_model_profile(settings).id == "openai-gpt-5-4"


def test_profile_url_preserves_gemini_generate_content_colon(tmp_path: Path) -> None:
    config_path = tmp_path / "config.yaml"
    config_path.write_text(
        f"""
models:
  chat_model: "gemini-3-1-pro"
  available_chat_models:
    - id: "gemini-3-1-pro"
      label: "Gemini 3.1 Pro Preview"
      provider: "gemini_generate_content"
      model: "gemini-3.1-pro-preview"
      base_url: "https://generativelanguage.googleapis.com/v1beta/models"
      chat_completions_path: "/gemini-3.1-pro-preview:generateContent"
      api_key_env: "GEMINI_API_KEY"
openai:
  api_key_env: "AI_GATEWAY_API_KEY"
paths:
  docs_dir: "{(tmp_path / "docs").as_posix()}"
  jars_dir: "{(tmp_path / "jars").as_posix()}"
  output_dir: "{(tmp_path / "output").as_posix()}"
  vector_store: "{(tmp_path / "output" / "vectors.sqlite3").as_posix()}"
""",
        encoding="utf-8",
    )
    settings = load_settings(config_path)
    profile = resolve_model_profile(settings, "gemini-3-1-pro")

    assert _profile_url(settings, profile) == "https://generativelanguage.googleapis.com/v1beta/models/gemini-3.1-pro-preview:generateContent"


def test_model_profile_can_use_yaml_api_key_without_exposing_it(monkeypatch, tmp_path: Path) -> None:
    config_path = tmp_path / "config.yaml"
    config_path.write_text(
        f"""
models:
  chat_model: "openai-gpt-4-1-nano-direct"
  available_chat_models:
    - id: "openai-gpt-4-1-nano-direct"
      label: "Direct OpenAI"
      provider: "openai_chat_completions"
      model: "gpt-4.1-nano"
      base_url: "https://api.openai.com"
      chat_completions_path: "/v1/chat/completions"
      api_key: "yaml-direct-key"
      api_key_env: "OPENAI_API_KEY"
openai:
  api_key_env: "AI_GATEWAY_API_KEY"
paths:
  docs_dir: "{(tmp_path / "docs").as_posix()}"
  jars_dir: "{(tmp_path / "jars").as_posix()}"
  output_dir: "{(tmp_path / "output").as_posix()}"
  vector_store: "{(tmp_path / "output" / "vectors.sqlite3").as_posix()}"
""",
        encoding="utf-8",
    )
    settings = load_settings(config_path)
    monkeypatch.delenv("OPENAI_API_KEY", raising=False)
    captured = {}

    class FakeResponse:
        def raise_for_status(self) -> None:
            return None

        def json(self) -> dict:
            return {"choices": [{"message": {"content": "ok"}}]}

    def fake_post(url, headers, json, timeout, verify=True):
        captured.update({"headers": headers})
        return FakeResponse()

    monkeypatch.setattr("src.agents.llm_client.requests.post", fake_post)

    assert request_text_completion(settings, "Say OK", model_id="openai-gpt-4-1-nano-direct") == "ok"
    assert captured["headers"]["Authorization"] == "Bearer yaml-direct-key"
    public_options = model_options_payload(settings)
    assert "api_key" not in public_options[0]
    assert "yaml-direct-key" not in str(public_options)


def test_missing_api_key_records_token_usage_row(monkeypatch, tmp_path: Path) -> None:
    config_path = tmp_path / "config.yaml"
    config_path.write_text(
        f"""
models:
  chat_model: "openai-gpt-4-1-nano-direct"
  available_chat_models:
    - id: "openai-gpt-4-1-nano-direct"
      label: "Direct OpenAI"
      provider: "openai_chat_completions"
      model: "gpt-4.1-nano"
      base_url: "https://api.openai.com"
      chat_completions_path: "/v1/chat/completions"
      api_key_env: "OPENAI_API_KEY"
openai:
  api_key_env: "AI_GATEWAY_API_KEY"
paths:
  docs_dir: "{(tmp_path / "docs").as_posix()}"
  jars_dir: "{(tmp_path / "jars").as_posix()}"
  output_dir: "{(tmp_path / "output").as_posix()}"
  vector_store: "{(tmp_path / "output" / "vectors.sqlite3").as_posix()}"
""",
        encoding="utf-8",
    )
    settings = load_settings(config_path)
    monkeypatch.delenv("OPENAI_API_KEY", raising=False)
    monkeypatch.delenv("AI_GATEWAY_API_KEY", raising=False)

    assert request_text_completion(settings, "Hello", model_id="openai-gpt-4-1-nano-direct", action="unit_missing_key") is None

    usage_path = tmp_path / "output" / "token_usage.xlsx"
    assert usage_path.exists()
    from openpyxl import load_workbook

    sheet = load_workbook(usage_path).active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[-1][1:5] == ("unit_missing_key", "openai_chat_completions", "gpt-4.1-nano", "missing_api_key")


def test_provider_http_error_is_surfaced_for_strict_design(monkeypatch, tmp_path: Path) -> None:
    config_path = tmp_path / "config.yaml"
    config_path.write_text(
        f"""
models:
  chat_model: "openai-gpt-4-1-nano-direct"
  available_chat_models:
    - id: "openai-gpt-4-1-nano-direct"
      label: "Direct OpenAI"
      provider: "openai_chat_completions"
      model: "gpt-4.1-nano"
      base_url: "https://api.openai.com"
      chat_completions_path: "/v1/chat/completions"
      api_key: "sk-unit-test-secret"
      api_key_env: "OPENAI_API_KEY"
      verify_ssl: false
openai:
  api_key_env: "AI_GATEWAY_API_KEY"
paths:
  docs_dir: "{(tmp_path / "docs").as_posix()}"
  jars_dir: "{(tmp_path / "jars").as_posix()}"
  output_dir: "{(tmp_path / "output").as_posix()}"
  vector_store: "{(tmp_path / "output" / "vectors.sqlite3").as_posix()}"
""",
        encoding="utf-8",
    )
    settings = load_settings(config_path)
    captured = {}

    class FakeResponse:
        status_code = 401
        text = '{"error":{"message":"Incorrect API key provided: sk-unit-test-secret","code":"invalid_api_key"}}'

        def json(self) -> dict:
            return {"error": {"message": "Incorrect API key provided: sk-unit-test-secret", "code": "invalid_api_key"}}

    def fake_post(url, headers, json, timeout, verify=True):
        captured["verify"] = verify
        return FakeResponse()

    monkeypatch.setattr("src.agents.llm_client.requests.post", fake_post)

    with pytest.raises(LLMRequestError) as error:
        request_text_completion(
            settings,
            "Return JSON only",
            model_id="openai-gpt-4-1-nano-direct",
            action="workflow_design",
            raise_on_error=True,
        )

    assert "Provider HTTP 401" in str(error.value)
    assert "invalid_api_key" in str(error.value)
    assert "sk-unit-test-secret" not in str(error.value)
    assert "sk-***" in str(error.value)
    assert captured["verify"] is False
