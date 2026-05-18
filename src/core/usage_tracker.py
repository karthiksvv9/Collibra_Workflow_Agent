from __future__ import annotations

import csv
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from src.core.config import Settings, settings


HEADERS = [
    "timestamp",
    "action",
    "provider",
    "model",
    "status",
    "prompt_tokens",
    "completion_tokens",
    "total_tokens",
    "prompt_chars",
    "completion_chars",
]


def ensure_usage_workbook(config: Settings = settings) -> Path:
    output_dir = config.paths.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = output_dir / "token_usage.xlsx"
    if xlsx_path.exists():
        return xlsx_path
    try:
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Token Usage"
        sheet.append(HEADERS)
        workbook.save(xlsx_path)
        return xlsx_path
    except Exception:
        csv_path = output_dir / "token_usage.csv"
        if not csv_path.exists():
            with csv_path.open("w", newline="", encoding="utf-8") as handle:
                csv.writer(handle).writerow(HEADERS)
        return csv_path


def record_usage(
    action: str,
    provider: str,
    model: str,
    prompt: str,
    completion: str,
    usage: dict[str, Any] | None = None,
    status: str = "ok",
    config: Settings = settings,
) -> Path:
    """Append token usage to output/token_usage.xlsx, with CSV fallback."""
    usage = usage or {}
    prompt_tokens = _usage_int(usage, "prompt_tokens", "input_tokens") or _estimate_tokens(prompt)
    completion_tokens = _usage_int(usage, "completion_tokens", "output_tokens") or _estimate_tokens(completion)
    total_tokens = _usage_int(usage, "total_tokens") or prompt_tokens + completion_tokens
    row = [
        datetime.now(timezone.utc).isoformat(),
        action,
        provider,
        model,
        status,
        prompt_tokens,
        completion_tokens,
        total_tokens,
        len(prompt or ""),
        len(completion or ""),
    ]
    output_dir = config.paths.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = ensure_usage_workbook(config)
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(xlsx_path)
        sheet = workbook.active
        sheet.append(row)
        workbook.save(xlsx_path)
        return xlsx_path
    except Exception:
        csv_path = output_dir / "token_usage.csv"
        write_header = not csv_path.exists()
        with csv_path.open("a", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            if write_header:
                writer.writerow(HEADERS)
            writer.writerow(row)
        return csv_path


def _usage_int(usage: dict[str, Any], *keys: str) -> int:
    for key in keys:
        value = usage.get(key)
        if isinstance(value, dict):
            continue
        try:
            return int(value)
        except (TypeError, ValueError):
            continue
    return 0


def _estimate_tokens(text: str) -> int:
    return max(1 if text else 0, int(len(text or "") / 4))
