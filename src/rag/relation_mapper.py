from __future__ import annotations

import re
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET


UUID_RE = re.compile(
    r"\b[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}\b"
)


@dataclass(slots=True)
class SemanticRelation:
    source: str
    target: str
    relation_type: str
    evidence: str
    source_file: str
    confidence: float = 0.75
    metadata: dict[str, Any] = field(default_factory=dict)


@dataclass(slots=True)
class RelationGraph:
    relations: list[SemanticRelation] = field(default_factory=list)
    uuid_index: dict[str, list[str]] = field(default_factory=dict)
    columns: dict[str, list[str]] = field(default_factory=dict)
    bpmn_nodes: dict[str, dict[str, Any]] = field(default_factory=dict)
    sequence_flows: list[dict[str, str]] = field(default_factory=list)

    def merge(self, other: "RelationGraph") -> "RelationGraph":
        self.relations.extend(other.relations)
        for key, values in other.uuid_index.items():
            self.uuid_index.setdefault(key, []).extend(value for value in values if value not in self.uuid_index.get(key, []))
        for key, values in other.columns.items():
            self.columns.setdefault(key, []).extend(value for value in values if value not in self.columns.get(key, []))
        self.bpmn_nodes.update(other.bpmn_nodes)
        self.sequence_flows.extend(other.sequence_flows)
        return self


class RelationMapper:
    def __init__(
        self,
        uuid_columns: list[str] | None = None,
        role_columns: list[str] | None = None,
        sample_rows: int = 100,
    ) -> None:
        self.uuid_columns = {value.lower() for value in (uuid_columns or [])}
        self.role_columns = {value.lower() for value in (role_columns or [])}
        self.sample_rows = sample_rows

    def map_path(self, path: str | Path) -> RelationGraph:
        file_path = Path(path)
        suffix = file_path.suffix.lower()
        if suffix == ".xlsx":
            return self._map_xlsx(file_path)
        if suffix in {".xml", ".bpmn"}:
            return self._map_xml(file_path)
        if suffix in {".app", ".form"}:
            return self._map_json_like(file_path)
        if suffix == ".zip":
            return self._map_zip(file_path)
        text = file_path.read_text(encoding="utf-8", errors="ignore")
        graph = RelationGraph()
        graph.uuid_index[str(file_path)] = sorted(set(UUID_RE.findall(text)))
        return graph

    def _map_xlsx(self, path: Path) -> RelationGraph:
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(path, data_only=True, read_only=True)
            graph = RelationGraph()
            for sheet in workbook.worksheets:
                rows = list(sheet.iter_rows(values_only=True))
                self._map_table_rows(path, sheet.title, rows, graph)
            return graph
        except Exception:
            return self._map_xlsx_zip(path)

    def _map_table_rows(
        self,
        path: Path,
        sheet_name: str,
        rows: list[tuple[Any, ...]] | list[list[Any]],
        graph: RelationGraph,
    ) -> None:
        if not rows:
            return
        headers = [normalise_header(value) for value in rows[0]]
        graph.columns[f"{path.name}:{sheet_name}"] = [header for header in headers if header]
        source_idx = _first_matching(headers, ["source", "source id", "source uuid", "source asset", "from"])
        target_idx = _first_matching(headers, ["target", "target id", "target uuid", "target asset", "to"])
        relation_idx = _first_matching(headers, ["relation", "relation type", "relation type id", "type", "predicate"])
        asset_idx = _first_matching(headers, ["asset", "asset id", "asset uuid", "name"])
        role_idx = _first_matching(headers, list(self.role_columns) + ["role", "responsibility", "owner", "steward"])

        for row_number, row in enumerate(rows[1 : self.sample_rows + 1], start=2):
            values = ["" if value is None else str(value).strip() for value in row]
            evidence = f"{path.name}:{sheet_name}:row {row_number}"
            for header, value in zip(headers, values, strict=False):
                if UUID_RE.search(value) or header in self.uuid_columns:
                    graph.uuid_index.setdefault(header or "uuid", [])
                    if value and value not in graph.uuid_index[header or "uuid"]:
                        graph.uuid_index[header or "uuid"].append(value)
            if source_idx is not None and target_idx is not None:
                source = _value_at(values, source_idx)
                target = _value_at(values, target_idx)
                relation_type = _value_at(values, relation_idx) if relation_idx is not None else "relatedTo"
                if source and target:
                    graph.relations.append(
                        SemanticRelation(
                            source=source,
                            target=target,
                            relation_type=relation_type or "relatedTo",
                            evidence=evidence,
                            source_file=str(path),
                            confidence=0.9,
                            metadata={"sheet": sheet_name, "row": row_number},
                        )
                    )
            if asset_idx is not None and role_idx is not None:
                asset = _value_at(values, asset_idx)
                role = _value_at(values, role_idx)
                if asset and role:
                    graph.relations.append(
                        SemanticRelation(
                            source=role,
                            target=asset,
                            relation_type="responsibleFor",
                            evidence=evidence,
                            source_file=str(path),
                            confidence=0.82,
                            metadata={"sheet": sheet_name, "row": row_number},
                        )
                    )

    def _map_xlsx_zip(self, path: Path) -> RelationGraph:
        from src.rag.documents import _read_shared_strings, _read_sheet_rows

        graph = RelationGraph()
        with zipfile.ZipFile(path) as archive:
            if "xl/workbook.xml" not in archive.namelist():
                return graph
            shared_strings = _read_shared_strings(archive)
            workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
            rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
            relation_targets = {
                rel.attrib["Id"]: rel.attrib["Target"]
                for rel in rels
                if rel.attrib.get("Target", "").startswith("worksheets/")
            }
            ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
            for sheet in workbook_root.iter():
                if not sheet.tag.endswith("}sheet"):
                    continue
                name = sheet.attrib.get("name", "Sheet")
                target = relation_targets.get(sheet.attrib.get(f"{{{ns}}}id", ""))
                if target:
                    rows = _read_sheet_rows(archive, f"xl/{target}", shared_strings)
                    self._map_table_rows(path, name, rows, graph)
        return graph

    def _map_xml(self, path: Path) -> RelationGraph:
        graph = RelationGraph()
        raw = path.read_text(encoding="utf-8", errors="ignore")
        graph.uuid_index[str(path)] = sorted(set(UUID_RE.findall(raw)))
        try:
            root = ET.fromstring(raw)
        except ET.ParseError:
            return graph
        for node in root.iter():
            tag = strip_ns(node.tag)
            node_id = node.attrib.get("id")
            if node_id and tag in BPMN_NODE_TAGS:
                graph.bpmn_nodes[node_id] = {
                    "id": node_id,
                    "type": tag,
                    "name": node.attrib.get("name", ""),
                    "documentation": _documentation(node),
                }
            if tag == "sequenceFlow":
                flow = {
                    "id": node.attrib.get("id", ""),
                    "sourceRef": node.attrib.get("sourceRef", ""),
                    "targetRef": node.attrib.get("targetRef", ""),
                    "name": node.attrib.get("name", ""),
                }
                graph.sequence_flows.append(flow)
                if flow["sourceRef"] and flow["targetRef"]:
                    graph.relations.append(
                        SemanticRelation(
                            source=flow["sourceRef"],
                            target=flow["targetRef"],
                            relation_type="sequenceFlow",
                            evidence=flow["id"] or str(path),
                            source_file=str(path),
                            confidence=1.0,
                            metadata=flow,
                        )
                    )
        return graph

    def _map_json_like(self, path: Path) -> RelationGraph:
        text = path.read_text(encoding="utf-8", errors="ignore")
        graph = RelationGraph()
        graph.uuid_index[str(path)] = sorted(set(UUID_RE.findall(text)))
        for match in re.finditer(r'"(?:id|key|name)"\s*:\s*"([^"]+)"', text):
            graph.columns.setdefault(path.name, []).append(match.group(1))
        return graph

    def _map_zip(self, path: Path) -> RelationGraph:
        graph = RelationGraph()
        with zipfile.ZipFile(path) as archive:
            for name in archive.namelist():
                suffix = Path(name).suffix.lower()
                if suffix not in {".bpmn", ".xml", ".form", ".app", ".json"}:
                    continue
                raw = archive.read(name).decode("utf-8", errors="ignore")
                temp_graph = RelationGraph()
                temp_graph.uuid_index[f"{path}:{name}"] = sorted(set(UUID_RE.findall(raw)))
                if suffix in {".bpmn", ".xml"}:
                    try:
                        root = ET.fromstring(raw)
                    except ET.ParseError:
                        graph.merge(temp_graph)
                        continue
                    for node in root.iter():
                        tag = strip_ns(node.tag)
                        node_id = node.attrib.get("id")
                        if node_id and tag in BPMN_NODE_TAGS:
                            temp_graph.bpmn_nodes[node_id] = {
                                "id": node_id,
                                "type": tag,
                                "name": node.attrib.get("name", ""),
                                "documentation": _documentation(node),
                            }
                        if tag == "sequenceFlow":
                            flow = {
                                "id": node.attrib.get("id", ""),
                                "sourceRef": node.attrib.get("sourceRef", ""),
                                "targetRef": node.attrib.get("targetRef", ""),
                                "name": node.attrib.get("name", ""),
                            }
                            temp_graph.sequence_flows.append(flow)
                            if flow["sourceRef"] and flow["targetRef"]:
                                temp_graph.relations.append(
                                    SemanticRelation(
                                        source=flow["sourceRef"],
                                        target=flow["targetRef"],
                                        relation_type="sequenceFlow",
                                        evidence=flow["id"] or f"{path}:{name}",
                                        source_file=f"{path}:{name}",
                                        confidence=1.0,
                                        metadata=flow,
                                    )
                                )
                graph.merge(temp_graph)
        return graph


BPMN_NODE_TAGS = {
    "process",
    "lane",
    "startEvent",
    "endEvent",
    "userTask",
    "scriptTask",
    "serviceTask",
    "exclusiveGateway",
    "parallelGateway",
    "subProcess",
    "callActivity",
}


def normalise_header(value: Any) -> str:
    return re.sub(r"\s+", " ", "" if value is None else str(value).strip().lower().replace("_", " "))


def strip_ns(value: str) -> str:
    return value.rsplit("}", 1)[-1] if "}" in value else value


def _first_matching(headers: list[str], candidates: list[str]) -> int | None:
    candidate_set = {normalise_header(candidate) for candidate in candidates if candidate}
    for index, header in enumerate(headers):
        if header in candidate_set:
            return index
    for index, header in enumerate(headers):
        if any(candidate in header for candidate in candidate_set):
            return index
    return None


def _value_at(values: list[str], index: int | None) -> str:
    if index is None or index >= len(values):
        return ""
    return values[index]


def _documentation(node: ET.Element) -> str:
    for child in node:
        if strip_ns(child.tag) == "documentation":
            return " ".join(child.itertext()).strip()
    return ""
